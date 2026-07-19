"""
Builds Student_Order_Tracker.xlsx: a personal business tracker for selling this
project as a mini/major project package, with two sheets:
  1. Orders Tracker — one row per student order, with revenue formulas
  2. Delivery Checklist — the standard per-order checklist, reusable every time
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

ARIAL = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="0B3D62")
HEADER_FONT = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=ARIAL, size=16, bold=True, color="0B3D62")
INPUT_FILL = PatternFill("solid", fgColor="FFF9E6")   # cells you fill in
FORMULA_FILL = PatternFill("solid", fgColor="F0F0F0")  # auto-computed, don't edit
FORMULA_FONT = Font(name=ARIAL, size=11, italic=True, color="555555")
NORMAL_FONT = Font(name=ARIAL, size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ============================================================
# Sheet 1: Orders Tracker
# ============================================================
ws = wb.active
ws.title = "Orders Tracker"

ws["A1"] = "Student Order Tracker"
ws["A1"].font = TITLE_FONT

# Summary block
ws["A3"] = "Total orders"
ws["B3"] = "=COUNTA(B9:B200)"
ws["A4"] = "Total revenue charged (\u20b9)"
ws["B4"] = "=SUM(H9:H200)"
ws["A5"] = "Total collected (\u20b9)"
ws["B5"] = "=SUM(I9:I200)"
ws["A6"] = "Total pending (\u20b9)"
ws["B6"] = "=B4-B5"
for r in range(3, 7):
    ws[f"A{r}"].font = Font(name=ARIAL, size=11, bold=True)
    ws[f"B{r}"].font = Font(name=ARIAL, size=11, bold=True, color="0F6E56")
    ws[f"B{r}"].number_format = '"\u20b9"#,##0;("\u20b9"#,##0)' if r != 3 else "0"

headers = [
    "Order ID", "Student Name", "College", "Contact", "Project Type",
    "Order Date", "Package Tier", "Price Charged (\u20b9)", "Amount Paid (\u20b9)",
    "Balance Due (\u20b9)", "Payment Status", "Delivery Status",
    "Submission Deadline", "Follow-up Date", "Uniqueness Notes", "General Notes",
]
HEADER_ROW = 8
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=HEADER_ROW, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BORDER

# Example row (realistic values) — row 9
example_row = 9
example_values = [
    "ORD-001", "Priya Sharma", "ABC Institute of Technology", "priya.s@example.com",
    "Mini Project", "2026-01-15", "Full Package", 10000, 5000,
    None,  # Balance Due — formula
    None,  # Payment Status — formula
    "In Progress", "2026-03-20", "2026-03-10",
    "Changed demo PDF domain to legal contracts instead of the generic sample",
    "Paid 50% upfront, remainder on delivery",
]
for i, v in enumerate(example_values, start=1):
    if v is not None:
        ws.cell(row=example_row, column=i, value=v)

# Formula cells for the example row + 40 blank rows below it (rows 9-49)
LAST_DATA_ROW = 49
for r in range(example_row, LAST_DATA_ROW + 1):
    ws[f"J{r}"] = f"=IF(H{r}=\"\",\"\",H{r}-I{r})"
    ws[f"K{r}"] = f"=IF(H{r}=\"\",\"\",IF(J{r}<=0,\"Paid\",IF(I{r}=0,\"Pending\",\"Partial\")))"
    ws[f"J{r}"].fill = FORMULA_FILL
    ws[f"K{r}"].fill = FORMULA_FILL
    ws[f"J{r}"].font = FORMULA_FONT
    ws[f"K{r}"].font = FORMULA_FONT
    ws[f"J{r}"].number_format = '"\u20b9"#,##0;("\u20b9"#,##0)'
    for col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 12, 13, 14, 15, 16]:
        cell = ws.cell(row=r, column=col)
        cell.font = NORMAL_FONT if r != example_row else NORMAL_FONT
        cell.fill = INPUT_FILL
        cell.border = BORDER
    ws[f"J{r}"].border = BORDER
    ws[f"K{r}"].border = BORDER
    ws[f"H{r}"].number_format = '"\u20b9"#,##0'
    ws[f"I{r}"].number_format = '"\u20b9"#,##0'

# Data validation dropdowns
dv_package = DataValidation(
    type="list",
    formula1='"Code only,Code + Explanation,Code + Documentation,Full Package"',
    allow_blank=True,
)
ws.add_data_validation(dv_package)
dv_package.add(f"G{example_row+1}:G{LAST_DATA_ROW}")

dv_project_type = DataValidation(type="list", formula1='"Mini Project,Major Project"', allow_blank=True)
ws.add_data_validation(dv_project_type)
dv_project_type.add(f"E{example_row+1}:E{LAST_DATA_ROW}")

dv_delivery = DataValidation(
    type="list", formula1='"Not Started,In Progress,Delivered"', allow_blank=True
)
ws.add_data_validation(dv_delivery)
dv_delivery.add(f"L{example_row+1}:L{LAST_DATA_ROW}")

# Column widths
widths = [10, 16, 22, 20, 13, 12, 16, 14, 14, 13, 13, 13, 15, 13, 28, 24]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Legend
legend_row = LAST_DATA_ROW + 3
ws.cell(row=legend_row, column=1, value="Legend").font = Font(name=ARIAL, bold=True)
ws.cell(row=legend_row + 1, column=1, value="Fill in yourself").fill = INPUT_FILL
ws.cell(row=legend_row + 1, column=1).font = NORMAL_FONT
ws.cell(row=legend_row + 2, column=1, value="Auto-calculated — don't edit").fill = FORMULA_FILL
ws.cell(row=legend_row + 2, column=1).font = FORMULA_FONT

ws.freeze_panes = f"A{HEADER_ROW + 1}"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

# ============================================================
# Sheet 2: Delivery Checklist
# ============================================================
ws2 = wb.create_sheet("Delivery Checklist")
ws2["A1"] = "Per-Order Delivery Checklist"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = "Same steps for every order — use the Notes column to jot order-specific details"
ws2["A2"].font = Font(name=ARIAL, size=10, italic=True, color="666666")

checklist_headers = ["#", "Step", "Done?", "Notes"]
CH_HEADER_ROW = 4
for i, h in enumerate(checklist_headers, start=1):
    c = ws2.cell(row=CH_HEADER_ROW, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

steps = [
    "Confirm package tier and price with the student before starting",
    "Collect student details: name, roll number, college, department, guide, HOD, academic year, submission month",
    "Update student_config.json with the student's details",
    "Run scripts/generate_all.sh to regenerate the report and presentation",
    "Replace Chapter 7 screenshot placeholders in the report with real app screenshots",
    "Replace the Demo slide placeholders in the presentation with real screenshots",
    "Open the report in Word: right-click the Table of Contents \u2192 Update Field \u2192 Update entire table",
    "Confirm whether the student's department requires IEEE or APA references specifically",
    "Vary something so this submission doesn't match another student's 1:1 (see notes column for ideas)",
    "Final read-through of report and slides for leftover placeholder text or brackets",
    "Package and send: report docx, presentation pptx, viva_question_bank.md",
    "Update Orders Tracker: payment status, delivery status, and dates",
    "Set a follow-up reminder around the student's viva date to check how it went",
]
for i, step in enumerate(steps):
    r = CH_HEADER_ROW + 1 + i
    ws2.cell(row=r, column=1, value=i + 1).font = NORMAL_FONT
    ws2.cell(row=r, column=2, value=step).font = NORMAL_FONT
    ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.cell(row=r, column=3).fill = INPUT_FILL
    ws2.cell(row=r, column=4).fill = INPUT_FILL
    for col in [1, 2, 3, 4]:
        ws2.cell(row=r, column=col).border = BORDER

dv_done = DataValidation(type="list", formula1='"No,Yes"', allow_blank=True)
ws2.add_data_validation(dv_done)
dv_done.add(f"C{CH_HEADER_ROW+1}:C{CH_HEADER_ROW+len(steps)}")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 70
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 30
for i in range(CH_HEADER_ROW + 1, CH_HEADER_ROW + 1 + len(steps)):
    ws2.row_dimensions[i].height = 32

ws2.freeze_panes = f"A{CH_HEADER_ROW + 1}"
ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Student_Order_Tracker.xlsx")
print("Saved Student_Order_Tracker.xlsx")
